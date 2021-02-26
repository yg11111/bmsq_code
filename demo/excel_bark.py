# /usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time: 2021/2/24 10:29
# @Author: yangguang
# @site: 
# @File: excel_bark.py
# @Software: PyCharm

import openpyxl


wb = openpyxl.load_workbook('mod启动失败数(每周).xlsx')
ws = wb.active
# s = ws.rows[1]
# print(type(s))
# c = ws['A1']
# print(c.value)
# c.value = 'gdgd'
# print(c.value)
# ws['A1'] = '应用id'
# s = ws.cell(row=1, column=1).value
# for cell_range in ws['A1':'C2']:
#     for x in range(3):
#         print(cell_range[x].value)
# print(ws['A'])
# for w in ws.iter_rows(min_row=2, max_row=2, min_col=2, max_col=3, values_only=True):
#     print(w)
# for w in ws.iter_cols(min_row=2, max_row=2, min_col=2, max_col=3):
#     print(w)
# print(ws.rows)
# print(ws.columns)
# print(ws.max_column)
# print(ws.max_row)
# for index in ws.values:
#     print(index)
# game_name_dict = dict()
# game_names = ws['A']
# for index in ws['B2:B43']:
#     game_name = index[0].value
#     game_name_row = index[0].row
#     print(game_name, game_name_row)
#     game_name_cell = 'D' + str(game_name_row)
#     game_name_dict[game_name] = game_name_cell
# print(game_name_dict)
# ws['A1'].value = "=SUM(1, 1)"
# ws["C1"] = "=SUM(1, 1)"
# # print(ws['C1'].value)
# wb.save("mod启动失败数(每周).xlsx")
cell = ws.rows
for index in cell:
    index[3] = 'GD'
