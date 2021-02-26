# -*- coding:utf-8 -*-
# @Time    : 2020/9/8 17:56
# @Author  : 年少风狂!
# @File    : packages_link.py
# @Software: PyCharm
from urllib import request
import time
import requests
import openpyxl
import urllib.request
import os

class Link_new(object):
    def __init__(self):
        pass

    def ask_package_link(self, url):
        response = urllib.request.urlopen(url)
        site = response.getheader("Content-Length")
        filesize = round(int(site) / (1024 * 1024), 2)
        code = response.status

        return filesize, code

    def sava_package_link(self):
        response = requests.head("https://dl2.bamenzhushou.com/apk/act1592812852841CD3H.apk")
        filesize = response.headers['Content-Length']
        headers = {
            'User-Agent': 'Mozilla/5.0 (Linux; U; Android 4.0.3; zh-cn; M032 Build/IML74K) '
                          'AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30',
            'Connection': 'keep-alive', }
        url = "https://dl2.bamenzhushou.com/apk/act1592812852841CD3H.apk"
        file = requests.get(url, headers=headers)
        apk_name = os.path.split(url)[1]
        print(apk_name)
        with open(f'F:\\chrome_download\\{apk_name}', "wb") as code:
            code.write(file.content)
            code.close()

    def main(self):
        wb = openpyxl.load_workbook("未知流量10-14.xlsx")
        booksheet = wb.get_sheet_by_name("能访问的包")
        line_number = booksheet.max_row
        for index in range(1339, line_number + 1):
            old_url = booksheet.cell(row=index, column=1).value
            new_url = booksheet.cell(row=index, column=2).value
            old_link = self.ask_package_link(old_url)
            new_link = self.ask_package_link(new_url)
            if (old_link[0] != new_link[0]) and (old_link[1] != new_link[1]) and \
                (os.path.basename(old_url) != os.path.basename(new_url)):
                print(f'{old_url}链接出现异常')
            print(f"第{index - 1}个链接已完成测试")


# if __name__ == '__main__':
#     link_new = Link_new()



link = 'https://dl2.bamengame.com/apk/20200911ReplaceMine/19.01.04snk4bm.apk'
link1 = 'https://dl2.bamengame.com/apk/20200911ReplaceMine/19.01.04snk4bm.apk'  # 平台上的链接
link2 = 'https://dl2.bamengame.com/apk/00a99f6f269e6ae64dfff78439710e97.apk' #
link3 = 'https://dl2.bamengame.com/apk/20200911ReplaceMine/act1594110169481JNhx.apk'  # 平台上没有该链接


for _ in range(3):
    requests.get(link3)
    # print(Link_new().ask_package_link(link3))
