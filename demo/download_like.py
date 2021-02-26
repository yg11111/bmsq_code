# -*- coding:utf-8 -*-
# @Time    : 2021/1/12 11:50
# @Author  : 年少风狂!
# @File    : download_like.py
# @Software: PyCharm
import requests
import json
import openpyxl


class BmSpiders(object):
    def __init__(self):
        self.start_url = "http://adminnew.bamenzhushou.com/"

    def get_token(self):  # 获取登录八门后台的token
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/84.0.4147.89 Safari/537.36'
        }
        url = f'{self.start_url}v1/admin/user/login/username'
        data = {"username": "yangguang", "password": "yangg079."}
        response = requests.post(url, data=data, headers=headers)
        html = response.text
        usertoken = json.loads(html)
        token = usertoken['content']['userToken']['token']
        with open('bm_channel-packages/token.txt', 'w') as f:
            f.write(token)

    def get_package(self):
        self.get_token()
        wb = openpyxl.load_workbook("应用下载日统计报表.xlsx")
        booksheet = wb.get_sheet_by_name("安装包")
        line_number = booksheet.max_row
        with open('bm_channel-packages/token.txt', 'r') as f:
            token = f.read()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/84.0.4147.89 Safari/537.36',
            'Authorization': token
        }
        wb1 = openpyxl.Workbook()
        sheet = wb1.create_sheet('Sheet1', 0)
        for index in range(1, line_number + 1):
            app_id = booksheet.cell(row=index, column=1).value
            app_name = booksheet.cell(row=index, column=2).value
            url = f'http://adminnew.bamenzhushou.com/api/app-admin-new/v1/info/getById?id={app_id}&includes='
            response = requests.get(url, headers=headers)
            html = response.text
            app_link = json.loads(html)
            link = app_link['content']['androidPackage']['downloadUrl']
            sheet.cell(index, 3, link)
        wb1.save('下载链接.xlsx')

    def get_downlad_like(self):
        wb = openpyxl.load_workbook("应用下载日统计报表.xlsx")
        booksheet = wb.get_sheet_by_name("安装包")
        line_number = booksheet.max_row
        for index in range(1, line_number + 1):
            app_link = booksheet.cell(row=index, column=3).value
            app_name = booksheet.cell(row=index, column=2).value
            response = requests.get(app_link)
            file = f'E:\\mod_game\\new\\{app_name}.apk'
            with open(file, "wb") as code:
                code.write(response.content)
            print(f'{index}')



bmspiders = BmSpiders()
bmspiders.get_downlad_like()