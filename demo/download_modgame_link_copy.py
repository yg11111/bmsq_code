# /usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time: 2021/2/19 10:38
# @Author: yangguang
# @site: 
# @File: download_modgame_link.py
# @Software: PyCharm
import requests
import json
import openpyxl
import time

class BmSpiders(object):
    def __init__(self):
        self.start_url = "http://adminnew.bamenzhushou.com/"

    def get_token(self):  # 获取登录八门后台的token
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/84.0.4147.89 Safari/537.36'
        }
        url = f'{self.start_url}v1/admin/user/login/username'
        data = {"username": "yangguang", "password": "yangg079."}
        response = requests.post(url, data=data, headers=headers)
        html = response.text
        usertoken = json.loads(html)
        token = usertoken['content']['userToken']['token']
        with open('token.txt', 'w') as f:
            f.write(token)

    def get_package(self):
        self.get_token()
        wb = openpyxl.load_workbook("mod启动失败数(每周) - 副本.xlsx")
        booksheet = wb.active
        # booksheet = wb.get_sheet_by_name("Sheet2")
        line_number = booksheet.max_row
        with open('token.txt', 'r') as f:
            token = f.read()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/84.0.4147.89 Safari/537.36',
            'Authorization': token
        }
        for index in range(2, line_number + 1):
            app_id = booksheet.cell(row=index, column=1).value
            app_name = booksheet.cell(row=index, column=2).value
            url = f'http://adminnew.bamenzhushou.com/api/app-admin-new/v1/info/getById?id={app_id}&includes='
            response = requests.get(url, headers=headers)
            html = response.text
            app_link = json.loads(html)
            link = app_link['content']['androidPackage']['downloadUrl']
            cell = 'C' + str(index)
            print(cell)
            booksheet[cell] = link
            # cell = booksheet.cell(index, 3, link).value
            wb.save("mod启动失败数(每周) - 副本.xlsx")
            # self.save_downlad_like(index, link)

    def save_downlad_like(self, index, link):
        wb = openpyxl.load_workbook("mod启动失败数(每周).xlsx")
        sheet = wb["Sheet2"]
        sheet.cell(index, 3, link)
        wb.save(r'F:\python_bmsq_code\mod启动失败数(每周).xlsx')

    def get_downlad_like(self):
        wb = openpyxl.load_workbook("mod启动失败数(每周).xlsx")
        booksheet = wb.get_sheet_by_name("Sheet2")
        line_number = booksheet.max_row
        for index in range(2, line_number + 1):
            app_link = booksheet.cell(row=index, column=3).value
            app_name = booksheet.cell(row=index, column=2).value
            response = requests.get(app_link)
            file = f'E:\\mod_game\\new1\\{app_name}.apk'
            with open(file, "wb") as code:
                code.write(response.content)
            print(f'{index-1}')


bmspiders = BmSpiders()
bmspiders.get_package()