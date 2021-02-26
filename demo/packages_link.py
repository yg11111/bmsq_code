# -*- coding:utf-8 -*-
# @Time    : 2020/9/8 17:56
# @Author  : 年少风狂!
# @File    : packages_link.py
# @Software: PyCharm
from urllib import request
import time
import requests
import openpyxl
import json
import os

class Flik(object):
    def __init__(self):
        self.start_url = "http://adminnew.bamenzhushou.com/"

    def get_login_token(self):  # 获取登录八门后台的token
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/84.0.4147.89 Safari/537.36'
        }
        url = f'{self.start_url}v1/admin/user/login/username'
        data = {"username": "yangguang", "password": "yangg079."}
        response = requests.post(url, data=data, headers=headers)
        html = response.text
        usertoken = json.loads(html)
        return usertoken['content']['userToken']['token']

    def get_packages_message(self, path, token):  # 根据包链接获取包名和MD5以及包大小
        try:
            url = f'http://adminnew.bamenzhushou.com/api/app-admin-new/v1/package/analysis-apk?url={path}'
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                              'AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/84.0.4147.89 Safari/537.36',
                'Authorization': token
            }
            response = requests.get(url, headers=headers)
            data = response.json()
            print(data)
            packagename = data["content"]["packageName"]
            signature = data["content"]["signature"]
            size = data["content"]["size"]

        except:
            print('token失效')
            new_token = self.get_login_token()
            url = f'http://adminnew.bamenzhushou.com/api/app-admin-new/v1/package/analysis-apk?url={path}'
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) '
                              'AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/84.0.4147.89 Safari/537.36',
                'Authorization': new_token
            }
            response = requests.get(url, headers=headers)
            data = response.json()
            packagename = data["content"]["packageName"]
            signature = data["content"]["signature"]
            size = data["content"]["size"]

        return packagename, signature, size

    def get_packages_link(self):  # 读取原始链接和更改链接的数据
        wb = openpyxl.load_workbook("2020-09-11可访问的链接(1).xlsx")
        sheet_name = wb.sheetnames
        # sheet = wb[sheet_name[0]]
        # table = wb.get_sheet_by_name('Sheet1')  # 工作簿的实例，按名称返回工作表
        active_sheet = wb.active   # 获取当前活跃的sheet,默认是第一个sheet
        # table = wb.get_sheet_by_name('Sheet1')  # 工作簿的实例，按名称返回工作表
        # max_row = table.max_row
        # sheets = wb.get_sheet_names()  # 从名称获取sheet
        # booksheet = wb.get_sheet_by_name(sheets[0])
        # rows = booksheet.rows  # 按行获取单元格(Cell对象) - 生成器
        # columns = booksheet.columns  # 按列获取单元格(Cell对象) - 生成器
        lang = active_sheet.max_row
        token = self.get_login_token()
        for x in range(47, lang+1):
            old_link = active_sheet[f'A{x}'].value  # 原始链接
            # active_sheet.cell()(row=1, column=1).value  # 通过坐标值来读取值
            new_link = active_sheet[f'B{x}'].value.split(':')
            new_links = f'{new_link[0]}s:{new_link[1]}'  # 更该链接
            old_package_message = self.get_packages_message(old_link, token)
            print(old_package_message)
            new_package_message = self.get_packages_message(new_links, token)
            if (old_package_message[0] != new_package_message[0]) and (old_package_message[1] != new_package_message[1]) \
                and (old_package_message[2] != new_package_message[2]) and (old_link.split('/')[4] != new_links.split('/')[5]):
                print(f'{old_link}原始链接出现异常')
            else:
                print(f'已完成第{x-1}个链接包的测试')


if __name__ == '__main__':
    link = Flik()
    link.get_packages_link()
