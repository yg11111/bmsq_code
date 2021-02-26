# -*- coding:utf-8 -*-
# @Time    : 2020/9/4 13:42
# @Author  : 年少风狂!
# @File    : taurus_selenium.py
# @Software: PyCharm


import time
import openpyxl
import requests
import json
import pickle
import datetime
from selenium import webdriver


class LoginSpider(object):
    def __init__(self):
        self.driver = webdriver.PhantomJS()

    def get_zhang_secret(self):
        wb = openpyxl.load_workbook("账密.xlsx")
        # sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        sheet = wb.get_sheet_by_name('Sheet1')
        username = sheet.cell(row=1, column=1).value
        password = sheet.cell(row=1, column=2).value

        return username, password

    def login(self):
        zhang_secret = self.get_zhang_secret()
        url = "http://login.bamenzhushou.com/login?service=http%3A%2F%2Ftaurus.admin.bamenzhushou.com%2Flogin"
        self.driver.get(url)
        self.driver.find_element_by_id("username").send_keys(f'{zhang_secret[0]}')
        self.driver.find_element_by_id("password").send_keys(f'{zhang_secret[1]}')
        self.driver.find_element_by_xpath('//div[@class="btm_b clearfix"]/button').submit()
        pickle.dump(self.driver.get_cookies(), open("cookies.pkl", "wb"))

    def select_statistical(self, user_name):
        start_url = "http://taurus.admin.bamenzhushou.com"
        url = "http://taurus.admin.bamenzhushou.com/bmbOrder/page?"
        headers = {
            'Host': 'taurus.admin.bamenzhushou.com',
            'Pragma': 'no-cache',
            'Referer': '{}/bmbOrder'.format(start_url),
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/'
                          '537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest'
        }
        today = datetime.date.today()
        params = {
            'limit': '20',
            'offset': '1',
            'userName': user_name,
            'orderNo': '',
            'payType': '',
            'payStatus': '0',
            'startCreateTime': today,
            'endCreateTime': f'{today + datetime.timedelta(days=1)}',
            'rechargeSource': '',
            'boxChannel': '',
            'rmbOrderNo': '',
            'channelOrderNo': '',
            'userId': '',
            'statisticsNo': '',
            'payChannel': '',
            'partnerId': '',
            'isRechargeFail': '0',
            'versionName': '',
            'platformId': '',
            'orderType': '',
            '_': str(int(time.time() * 1000))
        }
        req = requests.Session()
        cookies = pickle.load(open("cookies.pkl", "rb"))
        for cookie in cookies:
            req.cookies.set(cookie["name"], cookie["value"])
        response = req.get(url, params=params, headers=headers)
        data = json.loads(response.text)
        packages = data["content"][0]["packageName"]
        statisticsNo = data["content"][0]["statisticsNo"]

        return packages, statisticsNo

    def main(self, username):
        try:
            package_statisticsNo = self.select_statistical(username)
            self.driver.close()
            self.driver.quit()
        except:
            self.login()
            package_statisticsNo = self.select_statistical(username)
            self.driver.close()
            self.driver.quit()

        return package_statisticsNo
